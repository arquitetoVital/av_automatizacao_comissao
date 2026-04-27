"""
vendedores.py
─────────────
Leitura centralizada do arquivo vendedores.xlsx.
Substitui vendedores_sp.txt, vendedores_mg.txt, blacklist.txt e blacklist_clientes.txt.

Estrutura do xlsx:
  Aba VENDEDORES        → Nome | Filial (SP/MG) | Comissao (SIM/NÃO)
  Aba BLACKLIST_VENDEDORES → Nome
  Aba BLACKLIST_CLIENTES   → Termo

Uso:
    from vendedores import carregar_vendedores
    info = carregar_vendedores()

    info.filial("ISAQUE SANTOS")           # → "SP" | "MG" | None
    info.tem_comissao("ISAQUE SANTOS")     # → True | False
    info.na_blacklist_vendedor("HUGO")     # → True | False
    info.cliente_bloqueado("ACOS VITAL")  # → True | False
"""

import logging
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

log = logging.getLogger(__name__)

_XLSX = Path(__file__).parent / "vendedores.xlsx"


# ── Normalização ──────────────────────────────────────────────────────────────

def _norm(s: str) -> str:
    """Remove acentos, maiúsculo e strip."""
    s = unicodedata.normalize("NFD", str(s))
    return "".join(c for c in s if unicodedata.category(c) != "Mn").upper().strip()


def _nome_para_pasta(nome: str) -> str:
    """'HUGO DOS SANTOS' → 'HUGO_DOS_SANTOS'"""
    limpo = "".join(c for c in str(nome) if c not in r'\/:*?"<>|').strip()
    return limpo.replace(" ", "_")


# ── Estrutura de dados ────────────────────────────────────────────────────────

@dataclass
class InfoVendedores:
    """
    Contém todos os dados lidos do vendedores.xlsx.
    Chaves internas: nome normalizado no formato de pasta (underscores, maiúsculo).
    """
    # {chave_pasta → "SP" | "MG"}
    _filiais:     dict[str, str]       = field(default_factory=dict)
    # conjunto de chaves sem comissão
    _sem_comissao: set[str]            = field(default_factory=set)
    # blacklist de vendedores — exclui completamente
    _bl_vendedores: set[str]           = field(default_factory=set)
    # blacklist de clientes — termos normalizados para busca por substring
    _bl_clientes: list[str]            = field(default_factory=list)

    def _chave(self, nome_vendedor: str) -> str:
        return _nome_para_pasta(nome_vendedor).upper()

    # ── Consultas públicas ────────────────────────────────────────────────────

    def filial(self, nome_vendedor: str) -> str | None:
        """Retorna 'SP', 'MG' ou None se não cadastrado."""
        return self._filiais.get(self._chave(nome_vendedor))

    def tem_comissao(self, nome_vendedor: str) -> bool:
        """Retorna False se o vendedor está marcado com Comissao=NÃO."""
        return self._chave(nome_vendedor) not in self._sem_comissao

    def na_blacklist_vendedor(self, nome_vendedor: str) -> bool:
        """Retorna True se o vendedor deve ser completamente ignorado."""
        return self._chave(nome_vendedor) in self._bl_vendedores

    def cliente_bloqueado(self, nome_cliente: str) -> bool:
        """Retorna True se algum termo da blacklist é substring do nome do cliente."""
        nome_norm = _norm(nome_cliente)
        return any(termo in nome_norm for termo in self._bl_clientes)

    # ── Listas para uso no reports.py ────────────────────────────────────────

    def lista_sp(self) -> set[str]:
        """Conjunto de chaves (formato pasta) dos vendedores SP."""
        return {k for k, v in self._filiais.items() if v == "SP"}

    def lista_mg(self) -> set[str]:
        """Conjunto de chaves (formato pasta) dos vendedores MG."""
        return {k for k, v in self._filiais.items() if v == "MG"}


# ── Leitura do xlsx ───────────────────────────────────────────────────────────

def carregar_vendedores() -> InfoVendedores:
    """
    Lê vendedores.xlsx e retorna InfoVendedores preenchido.
    Chamada uma única vez por execução — resultado pode ser reutilizado.
    """
    info = InfoVendedores()

    if not _XLSX.exists():
        log.error(
            "  vendedores.xlsx nao encontrado em '%s'. "
            "Nenhum vendedor sera classificado.", _XLSX,
        )
        return info

    try:
        wb = openpyxl.load_workbook(_XLSX, read_only=True, data_only=True)
    except Exception as exc:
        log.error("  Erro ao abrir vendedores.xlsx: %s", exc)
        return info

    # ── Aba VENDEDORES ────────────────────────────────────────────────────────
    if "VENDEDORES" in wb.sheetnames:
        ws = wb["VENDEDORES"]
        total = sem_com = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            nome, filial, comissao = (row + (None, None, None))[:3]
            if not nome:
                continue
            chave  = _nome_para_pasta(str(nome)).upper()
            filial = str(filial or "").strip().upper()
            if filial in ("SP", "MG"):
                info._filiais[chave] = filial
                total += 1
            com = _norm(str(comissao or "SIM"))
            if com in ("NAO", "NÃO", "N", "NO", "FALSE", "0"):
                info._sem_comissao.add(chave)
                sem_com += 1
        log.info(
            "  VENDEDORES: %d cadastrados | %d sem comissao | %d SP | %d MG",
            total, sem_com,
            sum(1 for v in info._filiais.values() if v == "SP"),
            sum(1 for v in info._filiais.values() if v == "MG"),
        )
    else:
        log.warning("  Aba VENDEDORES nao encontrada em vendedores.xlsx.")

    # ── Aba BLACKLIST_VENDEDORES ──────────────────────────────────────────────
    if "BLACKLIST_VENDEDORES" in wb.sheetnames:
        ws = wb["BLACKLIST_VENDEDORES"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            nome = row[0]
            if nome:
                info._bl_vendedores.add(str(nome).strip().upper())
        log.info("  BLACKLIST_VENDEDORES: %d entradas.", len(info._bl_vendedores))
    else:
        log.warning("  Aba BLACKLIST_VENDEDORES nao encontrada.")

    # ── Aba BLACKLIST_CLIENTES ────────────────────────────────────────────────
    if "BLACKLIST_CLIENTES" in wb.sheetnames:
        ws = wb["BLACKLIST_CLIENTES"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            termo = row[0]
            if termo:
                info._bl_clientes.append(_norm(str(termo)))
        log.info("  BLACKLIST_CLIENTES: %d termo(s).", len(info._bl_clientes))
    else:
        log.warning("  Aba BLACKLIST_CLIENTES nao encontrada.")

    wb.close()
    return info
