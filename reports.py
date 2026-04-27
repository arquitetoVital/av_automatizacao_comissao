"""
reports.py
──────────
Geração dos arquivos Excel de comissão.
Recebe DataFrames prontos — sem acoplamento com services.

Melhoria #5: linhas com _em_erro=True são renderizadas em vermelho.
"""

import logging
import shutil
import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config
from utils import nome_para_pasta

log = logging.getLogger(__name__)

# ═══ LISTAS DE FILIAIS ═════════════════════════════════

def _carregar_lista(nome_arquivo: str) -> set[str]:
    caminho = Path(__file__).parent / nome_arquivo
    if not caminho.exists():
        log.warning("  %s nao encontrado.", nome_arquivo)
        return set()
    nomes: set[str] = set()
    for linha in caminho.read_text(encoding="utf-8").splitlines():
        linha = linha.strip()
        if linha and not linha.startswith("#"):
            nomes.add(linha.upper())
    log.info("  %s carregado: %d entradas.", nome_arquivo, len(nomes))
    return nomes


# ═══ LAYOUT DO EXCEL ═══════════════════════════════════

# Nomes de coluna internos (DataFrame) → rótulo exibido no Excel
# "Comissao_Vendedor_%" exibe como "Comissao Simulador"
# "Menor_Comissao_%"   exibe como "Comissao Real"
_RENOMEAR_HEADER: dict[str, str] = {
    "Comissao_Vendedor_%": "Comissao Simulador",
    "Comissao_Compras_%":  "Comissao Compras",
    "Menor_Comissao_%":    "Comissao Real",
    "Comissao_Definida_%": "Comissao Real",
}

COLUNAS_COORD = [
    "Data_Venda", "Numero_Pedido", "Nome_Cliente", "Nome_Vendedor",
    "Valor_Pedido", "Data_Nota_Fiscal", "Nota_Fiscal",
    "Valor_Faturado", "Valor_Pendente",
    "Comissao_Vendedor_%", "Comissao_Compras_%", "Menor_Comissao_%",
    "Valor_Comissao_Calculado",
    "Obs_Comissao",
]
LARGURAS_COORD = [14, 16, 32, 28, 14, 18, 14, 15, 15, 22, 18, 16, 22, 40]

COLUNAS_VENDOR = [
    "Data_Venda", "Numero_Pedido", "Nome_Cliente", "Nome_Vendedor",
    "Valor_Pedido", "Data_Nota_Fiscal", "Nota_Fiscal",
    "Valor_Faturado", "Valor_Pendente",
    "Comissao_Definida_%", "Valor_Comissao_Calculado",
    "Obs_Comissao",
]
LARGURAS_VENDOR = [14, 16, 32, 28, 14, 18, 14, 15, 15, 18, 22, 40]

COLUNAS_MOEDA = {"Valor_Pedido", "Valor_Faturado", "Valor_Pendente", "Valor_Comissao_Calculado"}
COLUNAS_PCT   = {"Comissao_Vendedor_%", "Comissao_Compras_%", "Menor_Comissao_%", "Comissao_Definida_%"}

COR_HEADER       = "1F4E79"
COR_LINHA_A      = "E2EFDA"   # linhas pares — verde claro  (Comissao Definida!)
COR_LINHA_B      = "F2F2F2"   # linhas ímpares — verde claro (Comissao Definida!)
COR_ERRO         = "FFCCCC"   # vermelho       — planilha rejeitada
COR_SEM_SIMULAD  = "FFF2CC"   # amarelo        — análise de compras pendente
COR_SEM_FATURA   = "EDEDED"   # cinza          — pedido não faturado
COR_FAB_INTERNA  = "DDEEFF"   # azul           — fabricação interna / simulador ausente
COR_LARANJA      = "FFD966"   # laranja        — adicione o simulador na pasta CUSTO

# Mapeamento obs → cor de fundo da linha.
# A cor verde (COR_LINHA_A/B alternado) é aplicada como fallback quando
# nenhuma obs especial corresponde — ou seja, "Comissao Definida!" e variantes.
_OBS_CORES: dict[str, str] = {
    "Ajuste a planilha de custo":                    COR_ERRO,
    "Analise de Compras pendente!":                  COR_SEM_SIMULAD,
    "Pedido ainda nao faturado":                     COR_SEM_FATURA,
    "Fabricacao interna / simulador ausente":        COR_FAB_INTERNA,
    "Adicione o simulador na pasta CUSTO":           COR_LARANJA,
    "Refaturamento":                                 COR_ERRO,    # vermelho — sem comissão
}

FMT_MOEDA = "R$ #,##0.00"
FMT_PCT   = "0.00%"


# ═══ ESCRITA DO EXCEL ══════════════════════════════════

def _escrever_excel(df: pd.DataFrame, caminho: Path, colunas: list, larguras: list) -> None:
    """
    Gera .xlsx em temp local e copia para o destino.
    Melhoria #5: linhas com _em_erro=True recebem fundo vermelho claro.
    Salva com timestamp se arquivo estiver bloqueado.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendas"

    font_h  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    fill_h  = PatternFill("solid", start_color=COR_HEADER)
    align_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borda   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    for col_idx, (col_name, larg) in enumerate(zip(colunas, larguras), start=1):
        header_label = _RENOMEAR_HEADER.get(col_name, col_name.replace("_", " "))
        cell = ws.cell(row=1, column=col_idx, value=header_label)
        cell.font      = font_h
        cell.fill      = fill_h
        cell.alignment = align_c
        cell.border    = borda
        ws.column_dimensions[get_column_letter(col_idx)].width = larg

    ws.row_dimensions[1].height = 28
    font_d = Font(name="Arial", size=9)

    # Garante que _em_erro está presente mas não inclui como coluna visível
    em_erro_col = "_em_erro"
    tem_em_erro = em_erro_col in df.columns

    df_out = df[colunas].copy()

    for r_idx, (_, row_series) in enumerate(df_out.iterrows(), start=2):
        # Verifica flag de erro a partir do df original (não do df_out filtrado)
        if tem_em_erro:
            linha_original_idx = r_idx - 2  # offset pelo header
            try:
                eh_erro = bool(df.iloc[linha_original_idx][em_erro_col])
            except (IndexError, KeyError):
                eh_erro = False
        else:
            eh_erro = False

        # Determina cor: obs especial tem prioridade sobre alternado verde/cinza
        obs_val = ""
        if "Obs_Comissao" in colunas:
            obs_idx = colunas.index("Obs_Comissao")
            obs_val = str(row_series.iloc[obs_idx]) if row_series.iloc[obs_idx] else ""

        if obs_val in _OBS_CORES:
            cor_fundo = _OBS_CORES[obs_val]
            eh_erro   = (cor_fundo == COR_ERRO)   # vermelho = negrito + texto escuro
        elif eh_erro:
            cor_fundo = COR_ERRO
        elif obs_val.startswith("Comissao Definida"):
            # Verde alternado para linhas com comissão confirmada
            cor_fundo = COR_LINHA_A if r_idx % 2 == 0 else COR_LINHA_B
        else:
            cor_fundo = COR_LINHA_A if r_idx % 2 == 0 else COR_LINHA_B

        fill_d = PatternFill("solid", start_color=cor_fundo)
        font_linha = Font(name="Arial", size=9, bold=eh_erro, color="990000" if eh_erro else "000000")

        for c_idx, value in enumerate(row_series, start=1):
            cell           = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font      = font_linha
            cell.fill      = fill_d
            cell.border    = borda
            cell.alignment = Alignment(vertical="center")

            col_name = colunas[c_idx - 1]
            if col_name in COLUNAS_MOEDA:
                cell.number_format = FMT_MOEDA
            elif col_name in COLUNAS_PCT:
                cell.number_format = FMT_PCT

    ws.freeze_panes = "A2"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)

    try:
        wb.save(tmp_path)
        caminho.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(tmp_path, caminho)
        log.info("  Arquivo salvo: %s  (%d linhas)", caminho, len(df_out))
    except PermissionError:
        ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = caminho.with_stem(f"{caminho.stem}_{ts}")
        shutil.copy2(tmp_path, fallback)
        log.warning("  '%s' bloqueado — salvo como: %s", caminho.name, fallback)
    finally:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except Exception:
                pass


# ═══ HELPERS ═══════════════════════════════════════════

def _nome_abreviado(nome_vendedor: str) -> str:
    """
    Retorna primeiro e último nome para uso no nome do arquivo.
    'ISAQUE RODRIGUES DOS SANTOS' -> 'ISAQUE_SANTOS'
    """
    partes = nome_vendedor.strip().split()
    if len(partes) <= 1:
        abrev = partes[0] if partes else nome_vendedor
    else:
        abrev = f"{partes[0]}_{partes[-1]}"
    return "".join(c for c in abrev if c not in r'\/:*?"<>|')


def _pasta_filial(nome_vendedor: str, lista_sp: set[str], lista_mg: set[str]) -> Path | None:
    chave = nome_para_pasta(nome_vendedor).upper()
    if chave in lista_sp:
        return config.PASTA_VENDEDOR_SP
    if chave in lista_mg:
        return config.PASTA_VENDEDOR_MG
    return None


# ═══ INTERFACE PÚBLICA ══════════════════════════════════

def gerar_relatorio_coordenador(df: pd.DataFrame) -> None:
    """Gera o relatório consolidado para o coordenador (com linhas vermelhas para erros)."""
    log.info("═══ Salvando planilha do coordenador ═══")
    caminho = config.PASTA_COORD / f"{config.MES_REF}_RELATORIO_GERAL_COMISSAO.xlsx"
    _escrever_excel(df, caminho, COLUNAS_COORD, LARGURAS_COORD)


def gerar_relatorio_analista(df: pd.DataFrame) -> None:
    """
    Gera cópia do relatório do coordenador na pasta da analista de vendas.
    Só executa se config.RELATORIO_ANALISTA_ATIVO == True.
    Mesmas colunas e formatação do relatório do coordenador.
    """
    if not config.RELATORIO_ANALISTA_ATIVO:
        log.info("  Relatório da analista desativado (RELATORIO_ANALISTA_ATIVO=false).")
        return
    log.info("═══ Salvando planilha da analista de vendas ═══")
    caminho = config.PASTA_ANALISTA / f"{config.MES_REF}_RELATORIO_GERAL_COMISSAO.xlsx"
    _escrever_excel(df, caminho, COLUNAS_COORD, LARGURAS_COORD)


def distribuir_para_vendedores(df: pd.DataFrame) -> None:
    """
    Gera uma planilha individual por vendedor na pasta de rede da filial correta.
    Melhoria #5: linhas em erro aparecem em vermelho também no relatório do vendedor.

    Pedidos com obs "Fabricacao interna" são ocultados do vendedor:
      - Obs substituída por "Adicione o simulador na pasta CUSTO"
      - Comissão zerada (vendedor não deve ver o cálculo interno)
    """
    log.info("═══ Distribuindo planilhas individuais ═══")

    df = df.copy()
    df["Comissao_Definida_%"] = df["Menor_Comissao_%"]

    # Oculta fabricação interna do vendedor — zera comissão e troca obs
    mask_fab = df["Obs_Comissao"] == "Fabricacao interna / simulador ausente"
    if mask_fab.any():
        df.loc[mask_fab, "Comissao_Definida_%"]        = 0.0
        df.loc[mask_fab, "Valor_Comissao_Calculado"]   = 0.0
        df.loc[mask_fab, "Obs_Comissao"]               = "Adicione o simulador na pasta CUSTO"
        log.info(
            "  %d linha(s) de fabricacao interna ocultadas nos relatorios dos vendedores.",
            int(mask_fab.sum()),
        )

    # Refaturamento: vendedor vê o pedido mas sem detalhes de comissão
    mask_refatur = df["Obs_Comissao"] == "Refaturamento"
    if mask_refatur.any():
        df.loc[mask_refatur, "Comissao_Definida_%"]      = 0.0
        df.loc[mask_refatur, "Valor_Comissao_Calculado"] = 0.0
        df.loc[mask_refatur, "Obs_Comissao"]             = "Pedido em analise"
        log.info(
            "  %d linha(s) de refaturamento ocultadas nos relatorios dos vendedores.",
            int(mask_refatur.sum()),
        )

    lista_sp = _carregar_lista("vendedores_sp.txt")
    lista_mg = _carregar_lista("vendedores_mg.txt")

    vendedores = df["Nome_Vendedor"].dropna().unique()
    log.info("  Vendedores: %d", len(vendedores))

    nao_classificados: list[str] = []

    for vendedor in sorted(vendedores):
        pasta_base = _pasta_filial(vendedor, lista_sp, lista_mg)

        if pasta_base is None:
            nao_classificados.append(vendedor)
            continue

        df_v       = df[df["Nome_Vendedor"] == vendedor].copy()
        nome_pasta = nome_para_pasta(vendedor)
        nome_arq   = _nome_abreviado(vendedor)
        destino    = (
            pasta_base
            / nome_pasta
            / config.MES_REF
            / f"{config.MES_REF}_COMISSAO_{nome_arq}.xlsx"
        )
        log.info("  '%s' -> %d pedidos  [%s]", vendedor, len(df_v), pasta_base.parts[-1])
        _escrever_excel(df_v, destino, COLUNAS_VENDOR, LARGURAS_VENDOR)

    if nao_classificados:
        log.warning(
            "  [AVISO] %d vendedor(es) sem filial definida — sem relatorio gerado:",
            len(nao_classificados),
        )
        for nome in nao_classificados:
            log.warning("    – %s", nome)

    log.info("  Distribuicao concluida.")
